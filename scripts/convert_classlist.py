#!/usr/bin/env python3
"""Refresh students.json from the live IB1 Google Sheet and local IB2 spreadsheet."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "students.json"
CACHE_DIR = ROOT / "tmp"
IB1_CACHE = CACHE_DIR / "ib1-class-list.xlsx"
IB2_XLSX = ROOT / "IB2 Class list 2025-2026.xlsx"

# Public IB1 class list. gid=0 is the first sheet.
IB1_SHEET_ID = "1pGFqN1XrL4T_NDr_a9fkAmuwMf5ob3e_0_f6_uFDC8c"
IB1_SHEET_GID = 0
IB1_EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{IB1_SHEET_ID}"
    f"/export?format=xlsx&gid={IB1_SHEET_GID}"
)

STUDENT_START_ROW = 8
BLOCKS = list("ABCDEFGH")


def strip_accents(text: str) -> str:
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def normalize_name(name: str) -> str:
    name = name.replace("\t", " ").replace("\n", " ")
    name = strip_accents(name)
    name = name.replace("’", "'").replace("`", "'")
    name = re.sub(r"[^A-Za-z0-9'\- ]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip().lower()
    return name


def tokens(name: str) -> set[str]:
    return {t for t in re.findall(r"[a-z0-9]+", normalize_name(name)) if t}


def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(
                min(prev[j] + 1, curr[j - 1] + 1, prev[j - 1] + (ca != cb))
            )
        prev = curr
    return prev[-1]


def display_name(raw: str) -> str:
    cleaned = re.sub(r"\s+", " ", raw.replace("\t", " ").replace("\n", " ")).strip()
    words = cleaned.split()
    if not words:
        return cleaned
    upperish = sum(1 for w in words if w.isupper() and len(w) > 1)
    if cleaned.isupper() or cleaned.islower() or upperish >= max(1, len(words) / 2):
        return cleaned.title()
    return cleaned


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", normalize_name(name)).strip("-")
    return slug or "student"


def last_similar(a: str, b: str) -> bool:
    la, lb = normalize_name(a).split()[-1], normalize_name(b).split()[-1]
    if la == lb:
        return True
    if la.startswith(lb) or lb.startswith(la):
        return min(len(la), len(lb)) >= 4
    return levenshtein(la, lb) <= 3 and min(len(la), len(lb)) >= 4


def should_merge(a: str, b: str) -> bool:
    if a == b:
        return True
    ta, tb = tokens(a), tokens(b)
    if len(ta) < 2 or len(tb) < 2:
        return False
    shorter, longer = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    if shorter.issubset(longer) and len(shorter) >= 2:
        return True
    jaccard = len(ta & tb) / len(ta | tb)
    na, nb = normalize_name(a).split(), normalize_name(b).split()
    first_similar = na[0] == nb[0] or levenshtein(na[0], nb[0]) <= 2
    if first_similar and last_similar(a, b) and len(ta & tb) >= 2:
        return True
    if na[0] == nb[0] and len(na) >= 3 and len(nb) >= 3 and na[1] == nb[1]:
        if levenshtein(na[-1], nb[-1]) <= 3:
            return True
    return jaccard >= 0.75


class UnionFind:
    def __init__(self, items: list[str]) -> None:
        self.parent = {x: x for x in items}

    def find(self, x: str) -> str:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: str, b: str) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def normalize_level(level: str) -> str:
    level = (level or "").strip().upper()
    mapping = {"TK": "TOK", "NM": "SL"}
    return mapping.get(level, level)


def normalize_subject(subject: str) -> str:
    subject = re.sub(r"\s+", " ", (subject or "").strip())
    aliases = {
        "Eng L&L": "English Lang & Lit",
        "English LangLit": "English Lang & Lit",
        "English LangLit SL": "English Lang & Lit",
        "English Lang & Lit": "English Lang & Lit",
        "Spanish Lang Lit": "Spanish Lang & Lit",
        "Spanish LangLit": "Spanish Lang & Lit",
        "Spanish LangLit SL": "Spanish Lang & Lit",
        "Spanish Ab Initio": "Spanish AB",
        "GloPo": "Global Politics",
        "Global Politics/Política Global": "Global Politics",
        "Math AA": "Math Analysis and Approaches",
        "Maths AA": "Math Analysis and Approaches",
        "Maths A&A": "Math Analysis and Approaches",
        "Maths A&I": "Math Applications and Interpretation",
        "Maths AI": "Math Applications and Interpretation",
        "Maths AI (in Spanish)": "Math Applications and Interpretation",
        "Mate A&I": "Math Applications and Interpretation",
        "Historia": "History",
        "SSST Lit": "SSST",
        "TDC": "TOK",
    }
    return aliases.get(subject, subject)


def cell_value(ws, merge_map, row: int, col: int):
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row, col).value


def build_merge_map(ws) -> dict[tuple[int, int], object]:
    merge_map: dict[tuple[int, int], object] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        value = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merge_map[(r, c)] = value
    return merge_map


def pick_canonical_raw(raws: list[str]) -> str:
    displays = [display_name(raw) for raw in raws]
    freq = Counter(displays)
    scored = []
    for cleaned in set(displays):
        toks = normalize_name(cleaned).split()
        unique = len(set(toks))
        dup = len(toks) - unique
        camel = 1 if re.search(r"[a-z][A-Z]", re.sub(r"\s+", "", cleaned)) else 0
        scored.append((freq[cleaned], unique, -dup, -camel, -len(toks), cleaned))
    scored.sort(reverse=True)
    return scored[0][5]


def convert_source(xlsx: Path, cohort: str, id_prefix: str) -> dict:
    if not xlsx.exists():
        raise SystemExit(f"Missing spreadsheet: {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    merge_map = build_merge_map(ws)

    classes: list[dict] = []
    for col in range(2, ws.max_column + 1):
        block = str(cell_value(ws, merge_map, 1, col) or "").strip().upper()
        subject = str(cell_value(ws, merge_map, 2, col) or "").strip()
        if block not in BLOCKS or not subject:
            continue
        teacher = str(cell_value(ws, merge_map, 4, col) or "").strip()
        room = cell_value(ws, merge_map, 5, col)
        room = "" if room is None else str(room).strip()
        level = normalize_level(str(cell_value(ws, merge_map, 3, col) or ""))
        names: list[str] = []
        for row in range(STUDENT_START_ROW, ws.max_row + 1):
            value = cell_value(ws, merge_map, row, col)
            if value is None:
                continue
            name = str(value).strip()
            if not name or name.isdigit():
                continue
            names.append(name)
        classes.append(
            {
                "block": block,
                "subject": normalize_subject(subject),
                "level": level,
                "teacher": teacher,
                "room": room,
                "students": names,
            }
        )

    raw_by_norm: dict[str, list[str]] = defaultdict(list)
    for cls in classes:
        for name in cls["students"]:
            key = normalize_name(name)
            if key and name not in raw_by_norm[key]:
                raw_by_norm[key].append(name)

    keys = list(raw_by_norm)
    uf = UnionFind(keys)
    for i, a in enumerate(keys):
        for b in keys[i + 1 :]:
            if should_merge(a, b):
                uf.union(a, b)

    clusters: dict[str, list[str]] = defaultdict(list)
    for key in keys:
        clusters[uf.find(key)].append(key)

    canonical_of: dict[str, str] = {}
    aliases_merged: list[dict] = []
    for _root, members in clusters.items():
        raws: list[str] = []
        for member in members:
            raws.extend(raw_by_norm[member])
        canonical_raw = pick_canonical_raw(raws)
        for member in members:
            canonical_of[member] = canonical_raw
        unique_norms = sorted(set(members))
        if len(unique_norms) > 1:
            aliases_merged.append(
                {
                    "canonical": canonical_raw,
                    "aliases": sorted(
                        {display_name(raw_by_norm[m][0]) for m in unique_norms}
                    ),
                }
            )

    students: dict[str, dict] = {}
    for cls in classes:
        entry = {
            "subject": cls["subject"],
            "level": cls["level"],
            "teacher": cls["teacher"],
            "room": cls["room"],
        }
        seen_in_class: set[str] = set()
        for raw in cls["students"]:
            key = normalize_name(raw)
            display = canonical_of[key]
            student_id = f"{id_prefix}{slugify(display)}"
            if student_id in seen_in_class:
                continue
            seen_in_class.add(student_id)
            student = students.setdefault(
                student_id,
                {
                    "id": student_id,
                    "name": display,
                    "cohort": cohort,
                    "blocks": {},
                },
            )
            block_list = student["blocks"].setdefault(cls["block"], [])
            fingerprint = (
                entry["subject"],
                entry["level"],
                entry["teacher"],
                entry["room"],
            )
            if any(
                (e["subject"], e["level"], e["teacher"], e["room"]) == fingerprint
                for e in block_list
            ):
                continue
            block_list.append(dict(entry))

    payload_students = []
    conflicts = []
    missing_blocks = []
    for student in sorted(students.values(), key=lambda s: s["name"].lower()):
        compact_blocks = {}
        for block, entries in student["blocks"].items():
            primary = dict(entries[0])
            if len(entries) > 1:
                primary["extras"] = entries[1:]
                conflicts.append(
                    {
                        "student": student["name"],
                        "block": block,
                        "classes": [
                            f"{e['subject']} {e['level']}" for e in entries
                        ],
                    }
                )
            compact_blocks[block] = primary
        present = set(compact_blocks)
        missing = [b for b in BLOCKS if b not in present]
        if missing:
            missing_blocks.append({"student": student["name"], "missing": missing})
        payload_students.append(
            {
                "id": student["id"],
                "name": student["name"],
                "cohort": cohort,
                "blocks": compact_blocks,
            }
        )

    return {
        "cohort": cohort,
        "source": xlsx.name,
        "students": payload_students,
        "class_columns": len(classes),
        "aliases_merged": aliases_merged,
        "conflicts": conflicts,
        "missing_blocks": missing_blocks,
    }


def print_cohort_report(result: dict) -> None:
    students = result["students"]
    print(f"\n{result['cohort']}  ({result['source']})")
    print(f"  Students: {len(students)}")
    print(f"  Class columns: {result['class_columns']}")
    print(f"  Aliases merged: {len(result['aliases_merged'])}")
    for item in result["aliases_merged"]:
        print(f"    - {item['canonical']}: {', '.join(item['aliases'])}")
    print(f"  Same-block conflicts: {len(result['conflicts'])}")
    for item in result["conflicts"]:
        print(
            f"    - {item['student']} block {item['block']}: "
            f"{', '.join(item['classes'])}"
        )
    notable_missing = [
        item for item in result["missing_blocks"] if len(item["missing"]) >= 2
    ]
    print(
        f"  Students missing 2+ lettered blocks: {len(notable_missing)} "
        f"(of {len(result['missing_blocks'])} with a study period)"
    )
    for item in notable_missing[:20]:
        print(f"    - {item['student']}: {', '.join(item['missing'])}")
    if len(notable_missing) > 20:
        print(f"    ... {len(notable_missing) - 20} more")


def download_sheet(url: str, dest: Path, label: str) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    partial = dest.with_suffix(dest.suffix + ".part")
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (compatible; uwccr-school-schedules)"},
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            data = response.read()
    except urllib.error.URLError as exc:
        raise SystemExit(f"Failed to download {label} class list: {exc}") from exc

    if not data.startswith(b"PK"):
        raise SystemExit(
            f"Downloaded {label} file is not an Excel workbook. "
            "Is the Google Sheet still shared with anyone who has the link?"
        )
    partial.write_bytes(data)
    partial.replace(dest)
    return dest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update IB1 from the public Google Sheet and IB2 from a local file."
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Use the last downloaded IB1 spreadsheet instead of fetching.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # --- IB1 ---
    if args.offline:
        if not IB1_CACHE.exists():
            raise SystemExit(f"No cached IB1 spreadsheet at {IB1_CACHE}")
        ib1_xlsx = IB1_CACHE
        print(f"Using cached IB1 spreadsheet: {ib1_xlsx.relative_to(ROOT)}")
    else:
        ib1_xlsx = download_sheet(IB1_EXPORT_URL, IB1_CACHE, "IB1")
        print(f"Downloaded IB1 class list ({ib1_xlsx.stat().st_size} bytes)")

    # --- IB2 (local file; sheet requires auth to download) ---
    if not IB2_XLSX.exists():
        raise SystemExit(
            f"Missing IB2 spreadsheet: {IB2_XLSX.name}\n"
            f"Export the IB2 class list from Google Sheets and save it as:\n"
            f"  {IB2_XLSX.relative_to(ROOT)}"
        )
    ib2_xlsx = IB2_XLSX
    print(f"Using IB2 spreadsheet: {ib2_xlsx.relative_to(ROOT)}")

    ib1 = convert_source(ib1_xlsx, "IB1", "")
    ib2 = convert_source(ib2_xlsx, "IB2", "ib2-")

    payload_students = [*ib1["students"], *ib2["students"]]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": f"{IB1_EXPORT_URL} + {IB2_XLSX.name}",
        "students": payload_students,
    }
    OUT.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print(f"\nWrote {OUT.relative_to(ROOT)}")
    print(f"Students: {len(payload_students)}")
    print_cohort_report(ib1)
    print_cohort_report(ib2)


if __name__ == "__main__":
    main()
