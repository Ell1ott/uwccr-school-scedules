#!/usr/bin/env python3
"""Parse the annual calendar spreadsheet into src/data/annualEvents.json."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "2026-2027 Final Annual Calendar.xlsx"
OUT = ROOT / "src" / "data" / "annualEvents.json"
SHEET = "Final Calendar 2026-27"

DAY_COLS = range(2, 9)  # B–H = Mon–Sun
COHORTS = ("IB1", "IB2")

RANGE_RE = re.compile(
    r"""
    (?P<h1>\d{1,2})
    (?:[:.](?P<m1>\d{2}))?
    \s*
    (?P<p1>a\.?m\.?|p\.?m\.?)?
    \s*[-–—/]\s*
    (?P<h2>\d{1,2})
    (?:[:.](?P<m2>\d{2}))?
    \s*
    (?P<p2>a\.?m\.?|p\.?m\.?)
    """,
    re.I | re.X,
)
SINGLE_RE = re.compile(
    r"(?<!\d)(?P<h>\d{1,2})(?:[:.](?P<m>\d{2}))?\s*(?P<p>a\.?m\.?|p\.?m\.?)",
    re.I,
)
BY_RE = re.compile(r"\bby\b", re.I)

SKIP_TITLE = re.compile(
    r"""
    students\ can\ only\ leave
    | \?\?
    | copyright
    | calendarpedia
    """,
    re.I | re.X,
)
THEME_RE = re.compile(r"^thematic week\s*#\s*\d+", re.I)
COMMUNITY_MEETING_RE = re.compile(r"community meeting\b", re.I)
NO_CLASS_RE = re.compile(
    r"""
    \bholiday\b
    | easter\ break
    | recharge\ days?
    | no\ classes
    | no\ lessons
    | community\ action\ week
    | student\ departures
    | campus\ closed
    | students\ on\ break
    """,
    re.I | re.X,
)
SAT_EXAM_RE = re.compile(r"\bsat exam", re.I)
COMMUNITY_DAY_RE = re.compile(r"community day\b", re.I)
EXAM_RE = re.compile(
    r"\bexams?\b|mock exams|official exams|end of semester exams",
    re.I,
)
STAFF_RE = re.compile(
    r"""
    faculty\ meeting
    | department\ meetings?
    | tok\ department
    | staff\ meeting
    | staff\ training
    | campus\ transition\ prep
    | \bmt\ extendido
    | \bmt\ meeting
    | board\ meeting
    | working\ days?\ for\ faculty
    | returning\ faculty
    | orientation/?planning\ for\ faculty
    | education\ team
    | ed\ team
    | webinar\ with
    | new\ staff\ orientation
    | all\ staff\ orientation
    | kick\ off\ session
    """,
    re.I | re.X,
)
OFFICE_RE = re.compile(r"office hours", re.I)
RESIDENCE_RE = re.compile(
    r"residence meeting|residences welcome|residence outing",
    re.I,
)
CAS_RE = re.compile(r"\bcas\b", re.I)
LIFE_RE = re.compile(r"life skills", re.I)
MENTOR_RE = re.compile(r"mentor group", re.I)
KEEP_UNTIMED = re.compile(
    r"""
    \bholiday\b
    | easter\ break
    | recharge
    | community\ action\ week
    | community\ day
    | \bexams?\b
    | graduation
    | student\ departures
    | uwc\ day
    | \bprom\b
    | first\ day\ of\ classes
    | arrival
    | student\ orientation
    """,
    re.I | re.X,
)


def fold(text: str) -> str:
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def tidy(text: str) -> str:
    text = text.replace("\xa0", " ").replace("´", "'").replace("’", "'")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip(" \n-")


def tidy_title(text: str) -> str:
    text = re.sub(r"\n+", " ", text)
    text = re.sub(r"\(\s*\)", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = text.strip(" \n\t-")
    text = re.sub(r"\(\s*$", "", text)
    text = re.sub(r"^\s*\)", "", text)
    text = re.sub(r"/+$", "", text)
    return text.strip(" \n\t-")


def preprocess_times(text: str) -> str:
    # "7:00-800pm" → "7:00-8:00pm"
    return re.sub(
        r"(?<=[-–/])\s*(\d)00(?=\s*(?:a\.?m\.?|p\.?m\.?))",
        r"\1:00",
        text,
        flags=re.I,
    )


def to_minutes(hours: int, minutes: int, pm: bool | None) -> int:
    hours = hours % 12
    if pm:
        hours += 12
    return hours * 60 + minutes


def parse_period(raw: str | None, fallback: bool | None) -> bool | None:
    if not raw:
        return fallback
    return raw.lower().startswith("p")


def fmt_hhmm(total: int) -> str:
    total = max(0, min(total, 23 * 60 + 59))
    return f"{total // 60:02d}:{total % 60:02d}"


def extract_range(text: str) -> tuple[str, str, str] | None:
    """Return (start, end, text_without_time) or None."""
    match = RANGE_RE.search(text)
    if not match:
        return None
    p2 = parse_period(match.group("p2"), None)
    p1 = parse_period(match.group("p1"), p2)
    if p2 is None and p1 is None:
        # Bare hours like 7-9: evening residential defaults to pm if >= 1 and <= 11
        h1 = int(match.group("h1"))
        p2 = h1 <= 11
        p1 = p2
    elif p2 is None:
        p2 = p1
    elif p1 is None:
        h1 = int(match.group("h1"))
        h2 = int(match.group("h2"))
        # 9:00-10:15am → both am; 12:00-2:00pm → first is noon/am-ish
        p1 = p2 if h1 != 12 else False
        if h1 > h2 and p2:
            p1 = False
    m1 = int(match.group("m1") or 0)
    m2 = int(match.group("m2") or 0)
    start = to_minutes(int(match.group("h1")), m1, p1)
    end = to_minutes(int(match.group("h2")), m2, p2)
    if end <= start:
        end += 12 * 60
    leftover = (text[: match.start()] + " " + text[match.end() :]).strip()
    leftover = tidy_title(leftover)
    return fmt_hhmm(start), fmt_hhmm(end), leftover


def extract_single(text: str) -> tuple[str, str, str] | None:
    if BY_RE.search(text):
        return None
    match = SINGLE_RE.search(text)
    if not match:
        return None
    pm = parse_period(match.group("p"), None)
    if pm is None:
        return None
    start = to_minutes(int(match.group("h")), int(match.group("m") or 0), pm)
    leftover = (text[: match.start()] + " " + text[match.end() :]).strip()
    leftover = tidy_title(leftover)
    return fmt_hhmm(start), fmt_hhmm(start + 60), leftover


def split_chunks(text: str) -> list[str]:
    parts = re.split(r"\n\s*\n", text)
    chunks: list[str] = []
    for part in parts:
        part = tidy(part)
        if not part:
            continue
        ranges = list(RANGE_RE.finditer(part))
        if len(ranges) >= 2:
            for i, match in enumerate(ranges):
                begin = 0 if i == 0 else ranges[i - 1].end()
                piece = tidy(part[begin : match.end()])
                if piece:
                    chunks.append(piece)
        else:
            chunks.append(part)
    return chunks or [text]


def guess_cohorts(text: str) -> list[str]:
    folded = fold(text).lower()
    ib1 = bool(re.search(r"\bib\s*1s?\b", folded))
    ib2 = bool(re.search(r"\bib\s*2s?\b", folded))
    if ib1 and ib2:
        return ["IB1", "IB2"]
    if ib1:
        return ["IB1"]
    if ib2:
        return ["IB2"]
    if re.search(r"all students", folded):
        return ["IB1", "IB2"]
    return ["IB1", "IB2"]


def guess_audience(text: str) -> str:
    if STAFF_RE.search(text) and not OFFICE_RE.search(text):
        if re.search(r"all students|ib\s*[12]", fold(text), re.I):
            return "both"
        return "staff"
    return "both"


def guess_kind(text: str) -> str:
    if COMMUNITY_MEETING_RE.search(text) or COMMUNITY_DAY_RE.search(text):
        return "community"
    if OFFICE_RE.search(text):
        return "office"
    if RESIDENCE_RE.search(text):
        return "residential"
    if NO_CLASS_RE.search(text) or THEME_RE.search(text):
        return "holiday"
    if EXAM_RE.search(text):
        return "activity"
    if CAS_RE.search(text) or LIFE_RE.search(text) or MENTOR_RE.search(text):
        return "activity"
    return "activity"


def guess_icon(kind: str, text: str) -> str:
    if kind == "community":
        return "users"
    if kind == "office":
        return "clock"
    if kind == "residential":
        return "moon" if "meeting" in text.lower() else "home"
    if kind == "holiday":
        return "info"
    if CAS_RE.search(text):
        return "mountain"
    if LIFE_RE.search(text):
        return "sparkles"
    if MENTOR_RE.search(text):
        return "users"
    if EXAM_RE.search(text):
        return "book-open"
    return "info"


def is_week_header(values: list[object]) -> bool:
    nums = [
        v
        for v in values
        if isinstance(v, (int, float)) and 1 <= int(v) <= 31
    ]
    return len(nums) >= 3


def merge_origin(ws, row: int, col: int) -> tuple[int, int, int, int]:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return rng.min_row, rng.min_col, rng.max_row, rng.max_col
        return row, col, row, col
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col, rng.max_row, rng.max_col
    return row, col, row, col


def cell_value(ws, row: int, col: int) -> object:
    min_r, min_c, _, _ = merge_origin(ws, row, col)
    return ws.cell(min_r, min_c).value


def infer_week_dates(
    numbers: list[int | None],
    year: int,
    month: int,
    last_day: int,
) -> tuple[list[date | None], int, int, int]:
    dates: list[date | None] = [None] * 7
    for i, n in enumerate(numbers):
        if n is None:
            continue
        if last_day and n < last_day:
            month += 1
            if month > 12:
                month = 1
                year += 1
        last_day = n
        dates[i] = date(year, month, n)
    known = [(i, d) for i, d in enumerate(dates) if d]
    if known:
        origin_i, origin_d = known[0]
        filled = [
            origin_d + timedelta(days=i - origin_i) for i in range(7)
        ]
        return filled, year, month, last_day
    return dates, year, month, last_day


def classify_event(title: str, start: str | None, end: str | None) -> dict:
    community_meeting = bool(COMMUNITY_MEETING_RE.search(title))
    no_classes = bool(NO_CLASS_RE.search(title) or COMMUNITY_DAY_RE.search(title))
    if EXAM_RE.search(title) and not SAT_EXAM_RE.search(title) and not start:
        no_classes = True
    if SAT_EXAM_RE.search(title):
        no_classes = False
    kind = guess_kind(title)
    if community_meeting:
        kind = "community"
        no_classes = False
    if no_classes and kind == "activity" and not EXAM_RE.search(title):
        kind = "holiday"
    return {
        "start": start,
        "end": end,
        "title": title,
        "kind": kind,
        "cohorts": guess_cohorts(title),
        "audience": guess_audience(title),
        "icon": guess_icon(kind, title),
        "noClasses": no_classes,
        "communityMeeting": community_meeting,
        "allDay": start is None,
    }


def parse_cell_text(raw: str) -> tuple[list[dict], list[str]]:
    text = preprocess_times(tidy(raw))
    skipped: list[str] = []
    events: list[dict] = []
    if not text or SKIP_TITLE.search(text):
        if text:
            skipped.append(text)
        return events, skipped

    if THEME_RE.search(text) and "event" not in text.lower():
        return events, skipped

    for chunk in split_chunks(text):
        chunk = preprocess_times(chunk)
        if SKIP_TITLE.search(chunk):
            skipped.append(chunk)
            continue
        if BY_RE.search(chunk) and not RANGE_RE.search(chunk):
            skipped.append(chunk)
            continue
        timed = extract_range(chunk)
        single = None if timed else extract_single(chunk)
        if timed:
            start, end, leftover = timed
            title = tidy_title(leftover or chunk) or tidy_title(chunk)
            events.append(classify_event(title, start, end))
        elif single:
            start, end, leftover = single
            title = tidy_title(leftover or chunk) or tidy_title(chunk)
            events.append(classify_event(title, start, end))
        elif KEEP_UNTIMED.search(chunk) or NO_CLASS_RE.search(chunk):
            events.append(classify_event(tidy_title(chunk), None, None))
        else:
            skipped.append(chunk)
    return events, skipped


def load_weeks(ws) -> list[dict]:
    max_row = min(ws.max_row or 1, 200)
    header_rows: list[int] = []
    for row in range(1, max_row + 1):
        values = [cell_value(ws, row, col) for col in DAY_COLS]
        if is_week_header(values):
            header_rows.append(row)

    year, month, last_day = 2026, 7, 0
    weeks: list[dict] = []
    for idx, row in enumerate(header_rows):
        next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else max_row + 1
        numbers: list[int | None] = []
        for col in DAY_COLS:
            v = cell_value(ws, row, col)
            if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
                numbers.append(int(v))
            else:
                numbers.append(None)
        dates, year, month, last_day = infer_week_dates(
            numbers, year, month, last_day
        )
        weeks.append(
            {
                "header_row": row,
                "end_row": next_header,
                "dates": dates,
            }
        )
    return weeks


def parse_calendar(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    weeks = load_weeks(ws)
    seen_origins: set[tuple[int, int]] = set()
    events: list[dict] = []
    themes: list[dict] = []
    skipped: list[dict] = []
    first_day: date | None = None

    for week in weeks:
        dates: list[date] = week["dates"]
        for row in range(week["header_row"], week["end_row"]):
            for col in DAY_COLS:
                min_r, min_c, max_r, max_c = merge_origin(ws, row, col)
                origin = (min_r, min_c)
                if origin in seen_origins:
                    continue
                value = ws.cell(min_r, min_c).value
                if value is None or isinstance(value, (int, float, time, datetime)):
                    continue
                if not isinstance(value, str):
                    continue
                text = tidy(value)
                if not text:
                    continue
                seen_origins.add(origin)
                covered: list[date] = []
                for c in range(min_c, max_c + 1):
                    idx = c - 2
                    if 0 <= idx < 7 and dates[idx]:
                        covered.append(dates[idx])
                if not covered:
                    continue
                if THEME_RE.search(text) and "event" not in text.lower():
                    themes.append(
                        {
                            "start": covered[0].isoformat(),
                            "end": covered[-1].isoformat(),
                            "title": tidy_title(text),
                        }
                    )
                    continue
                parsed, skip_bits = parse_cell_text(text)
                for bit in skip_bits:
                    skipped.append(
                        {
                            "dates": [d.isoformat() for d in covered],
                            "text": bit,
                        }
                    )
                for item in parsed:
                    if re.search(r"first day of classes", item["title"], re.I):
                        if first_day is None:
                            first_day = covered[0]
                        continue
                    if item["allDay"]:
                        item["start"] = "07:30"
                        item["end"] = "13:20"
                    for day in covered:
                        events.append(
                            {
                                "date": day.isoformat(),
                                **item,
                            }
                        )

    events.sort(key=lambda e: (e["date"], e.get("start") or "", e["title"]))
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": path.name,
        "firstDayOfClasses": (first_day or date(2026, 8, 19)).isoformat(),
        "themes": themes,
        "events": events,
        "skipped": skipped,
    }


def summarize(payload: dict) -> None:
    events = payload["events"]
    print(f"Events: {len(events)}")
    print(f"Themes: {len(payload['themes'])}")
    print(f"Skipped: {len(payload['skipped'])}")
    print(f"First day of classes: {payload['firstDayOfClasses']}")
    by_kind: dict[str, int] = {}
    community = []
    no_time = []
    for event in events:
        by_kind[event["kind"]] = by_kind.get(event["kind"], 0) + 1
        if event["communityMeeting"]:
            community.append(event)
        if event["allDay"]:
            no_time.append(event)
    print("By kind:", by_kind)
    print("\nCommunity meetings:")
    for event in community:
        print(
            f"  {event['date']} {event.get('start')}–{event.get('end')}  {event['title']}  {event['cohorts']}"
        )
    print("\nAll-day / untimed kept:")
    for event in no_time:
        print(
            f"  {event['date']}  {event['title']!r}  {event['cohorts']}  noClasses={event['noClasses']} audience={event['audience']}"
        )
    print("\nSkipped (review):")
    for item in payload["skipped"]:
        print(f"  {item['dates'][0]}  {item['text']!r}")


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX.name}")
    payload = parse_calendar(XLSX)
    skipped = payload.pop("skipped")
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {OUT.relative_to(ROOT)}")
    payload["skipped"] = skipped
    summarize(payload)


if __name__ == "__main__":
    main()
